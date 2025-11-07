"""
Excel 문서 파서
"""
from openpyxl import load_workbook
from datetime import datetime
from typing import List

from app.parsers.base import DocumentParser, ParsedDocument, DocumentMetadata, ContentChunk


class ExcelParser(DocumentParser):
    """Excel 문서 파서"""
    
    def parse(self, file_path: str) -> ParsedDocument:
        """Excel 문서 파싱"""
        wb = load_workbook(file_path, data_only=True)
        
        full_text = ""
        chunks = []
        structure = {
            "sheets": [],
            "tables": []
        }
        
        # 시트별 처리
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_text = f"시트: {sheet_name}\n"
            
            # 셀 데이터 추출
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    sheet_text += row_text + "\n"
            
            full_text += sheet_text + "\n"
            structure["sheets"].append({
                "name": sheet_name,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column
            })
        
        # 메타데이터 추출
        metadata = self.extract_metadata(file_path)
        metadata.word_count = len(full_text.split())
        
        # 청크 분할
        chunks = self.chunk_document(full_text)
        
        return ParsedDocument(
            filename=file_path.split("/")[-1],
            file_type="xlsx",
            metadata=metadata,
            chunks=chunks,
            full_text=full_text,
            structure=structure
        )
    
    def extract_metadata(self, file_path: str) -> DocumentMetadata:
        """Excel 메타데이터 추출"""
        metadata = DocumentMetadata()
        
        try:
            wb = load_workbook(file_path)
            props = wb.properties
            
            metadata.title = props.title or ""
            metadata.author = props.creator or ""
            
            if props.created:
                metadata.created_date = props.created
            if props.modified:
                metadata.modified_date = props.modified
        except Exception as e:
            print(f"메타데이터 추출 오류: {e}")
        
        return metadata

